"""
DIGEMID Modificatorias Scraper
Extrae Modificaciones al Registro Sanitario de DIGEMID y exporta Excel.
"""

import os, re, json, time
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import fitz  # pymupdf

# ── Constantes ─────────────────────────────────────────────────────────────────
BASE_URL  = "https://www.digemid.minsa.gob.pe"
LIST_URL  = f"{BASE_URL}/webDigemid/publicaciones/alertas-modificaciones/modificaciones/"
PAGE_URL  = LIST_URL + "page/{page}/"
PDF_DELAY = 12  # segundos entre descargas PDF

# ── Sesión con headers reales ──────────────────────────────────────────────────
_session = requests.Session()
_HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-PE,es;q=0.9,en-US;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer":         "https://www.digemid.minsa.gob.pe/",
}
try:
    _session.get(BASE_URL, headers=_HEADERS, timeout=30)
except Exception:
    pass


def _get_con_reintentos(url: str, timeout: int = 30, reintentos: int = 3, espera: float = 5):
    """GET con reintentos y backoff. Antes solo la descarga de PDF tenía esto;
    el listado, las páginas y el detalle de cada post fallaban en silencio ante
    un timeout/403/429 puntual (posible bloqueo WAF de datacenter). Devuelve el
    Response o None si se agotan los reintentos."""
    for intento in range(1, reintentos + 1):
        try:
            r = _session.get(url, headers=_HEADERS, timeout=timeout)
            if r.status_code == 200:
                return r
            if r.status_code in (403, 429, 503):
                print(f"    [{r.status_code}] {url[:70]}... reintento {intento}/{reintentos}")
                time.sleep(espera * intento)
                continue
            return r  # ej. 404 — no tiene sentido reintentar
        except Exception as e:
            print(f"    [error] {url[:70]}... reintento {intento}/{reintentos}: {e}")
            time.sleep(espera)
    return None

# ── Glosario de tiempos regulatorios (D.S. 016-2011-SA) ──────────────────────
# ══════════════════════════════════════════════════════════════════════════════
# TEXTOS ESTANDARIZADOS ConkoSafe IA — Modificatorias por seguridad
# Estos tres textos son fijos por decisión de Gestión de Riesgos: no se dejan
# al criterio del motor de análisis (Claude o heurístico) para que el reporte
# sea homogéneo entre corridas y auditable.
# ══════════════════════════════════════════════════════════════════════════════
TITULO_REPORTE = "ConkoSafe IA — Modificatorias por seguridad DIGEMID"

ACCION_SEGURIDAD = ("Evaluar impacto y comunicar posición a DIGEMID en 15 días hábiles "
                    "- Gestión de Riesgos")
TIEMPOS_SEGURIDAD = ("15 días hábiles para evaluar e informar al DIGEMID, "
                     "finalizar proceso según tiempos de RD y evaluar efectividad")
BASE_LEGAL_SEGURIDAD = "Art. 6,13,36,37,151  D.S. 016-2011-SA / ICH E2C"

# Tipos que este reporte trata como "de seguridad": reciben los textos fijos.
TIPOS_SEGURIDAD = {
    "ACTUALIZACION SEGURIDAD",
    "REACCION ADVERSA",
    "CONTRAINDICACION",
    "ADVERTENCIA",
    "SUSPENSION",
    "CANCELACION",
}

GLOSARIO_TIEMPOS = [
    ("FICHA TECNICA",          "60 días hábiles desde notificación DIGEMID",                  "Art. 58 D.S. 016-2011-SA"),
    ("ROTULADO",               "90 días hábiles para implementar cambio de rotulado",          "Art. 62 D.S. 016-2011-SA"),
    ("ACTUALIZACION SEGURIDAD", TIEMPOS_SEGURIDAD,                                             BASE_LEGAL_SEGURIDAD),
    ("REACCION ADVERSA",       "15 días hábiles (RAM grave) / 30 días (no grave)",             "Art. 55 D.S. 016-2011-SA"),
    ("CONTRAINDICACION",       "30 días hábiles para actualizar FT y rotulado",                "Art. 58 D.S. 016-2011-SA"),
    ("ADVERTENCIA",            "30 días hábiles para actualizar FT y material informativo",    "Art. 58 D.S. 016-2011-SA"),
    ("INDICACION",             "Requiere nueva evaluación; plazo según tipo de modificación",  "Art. 46 D.S. 016-2011-SA"),
    ("INTERACCION",            "30 días hábiles para actualizar ficha técnica",                "Art. 58 D.S. 016-2011-SA"),
    ("POSOLOGIA",              "60 días hábiles para actualizar FT y material de prescripción","Art. 58 D.S. 016-2011-SA"),
    ("INSCRIPCION",            "Trámite de registro: 90 días hábiles (D.S. 016)",              "Art. 14 D.S. 016-2011-SA"),
    ("SUSPENSION",             "Acción inmediata — retiro cautelar mientras dure la suspensión","Art. 78 D.S. 016-2011-SA"),
    ("CANCELACION",            "Retiro definitivo del mercado; informar en 5 días hábiles",    "Art. 80 D.S. 016-2011-SA"),
    ("GENERAL",                "Revisar comunicado; plazo según tipo específico indicado",     "D.S. 016-2011-SA / D.S. 020-2023-SA"),
]

URGENCIA_MAP = {
    "SUSPENSION":           "INMEDIATA",
    "CANCELACION":          "INMEDIATA",
    "ACTUALIZACION SEGURIDAD": "PREVENTIVA",
    "REACCION ADVERSA":     "PREVENTIVA",
    "CONTRAINDICACION":     "PREVENTIVA",
    "ADVERTENCIA":          "PREVENTIVA",
    "FICHA TECNICA":        "INFORMATIVA",
    "ROTULADO":             "INFORMATIVA",
    "INDICACION":           "INFORMATIVA",
    "INTERACCION":          "INFORMATIVA",
    "POSOLOGIA":            "INFORMATIVA",
    "INSCRIPCION":          "INFORMATIVA",
    "GENERAL":              "INFORMATIVA",
}

PLAZO_MAP = {g[0]: g[1] for g in GLOSARIO_TIEMPOS}
BASE_MAP  = {g[0]: g[2] for g in GLOSARIO_TIEMPOS}

# ── Prompt Claude API ──────────────────────────────────────────────────────────
_PROMPT = """\
Eres un experto en regulación farmacéutica peruana.
Analiza el texto de una Modificación al Registro Sanitario publicada por DIGEMID y extrae:
Responde SOLO con JSON válido, sin texto adicional ni bloques markdown:

{
  "tipo_modificacion": "FICHA TECNICA|ROTULADO|ACTUALIZACION SEGURIDAD|REACCION ADVERSA|CONTRAINDICACION|ADVERTENCIA|INDICACION|INTERACCION|POSOLOGIA|INSCRIPCION|SUSPENSION|CANCELACION|GENERAL",
  "principio_activo": "IFA principal",
  "titular_rs": "Laboratorio o empresa titular",
  "accion_requerida": "Qué debe hacer el Titular del Registro Sanitario",
  "resumen": "1-2 oraciones resumiendo el cambio"
}

TEXTO:
{texto}
"""

# ══════════════════════════════════════════════════════════════════════════════
# MOTORES DE ANÁLISIS
# ══════════════════════════════════════════════════════════════════════════════
def _analizar_claude(texto: str) -> dict | None:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return None
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model=os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6"),
            max_tokens=600,
            messages=[{"role": "user", "content": _PROMPT.replace("{texto}", texto[:8000])}]
        )
        raw = msg.content[0].text.strip()
        raw = re.sub(r"^```json\s*|^```\s*|\s*```$", "", raw, flags=re.MULTILINE).strip()
        return json.loads(raw)
    except Exception as e:
        print(f"    [Claude API] {e}")
        return None


def _tipo_desde_titulo(titulo: str, producto: str = "") -> str:
    up = (titulo + " " + (producto or "")).upper()
    if any(k in up for k in ["SUSPENS", "SUSPENDER"]):
        return "SUSPENSION"
    if any(k in up for k in ["CANCEL", "BAJA DEL REGISTRO"]):
        return "CANCELACION"
    if any(k in up for k in ["INSCRIPCION", "INSCRIPCIÓN", "REINSCRIPCION"]):
        return "INSCRIPCION"
    if any(k in up for k in ["REACCION ADVERSA", "RAM", "REACCIÓN ADVERSA", "EFECTO ADVERSO"]):
        return "REACCION ADVERSA"
    if any(k in up for k in ["CONTRAINDICACION", "CONTRAINDICACIÓN"]):
        return "CONTRAINDICACION"
    if any(k in up for k in ["ADVERTENCIA", "PRECAUCION", "PRECAUCIÓN"]):
        return "ADVERTENCIA"
    if any(k in up for k in ["SEGURIDAD", "SAFETY", "FARMACOVIGILANCIA", "INFORME SOBRE SEGURIDAD"]):
        return "ACTUALIZACION SEGURIDAD"
    if any(k in up for k in ["ROTULADO", "ETIQUETA", "PROSPECTO", "INSERTO"]):
        return "ROTULADO"
    if any(k in up for k in ["POSOLOGIA", "DOSIFICACION", "DOSIS", "POSOLOGÍA"]):
        return "POSOLOGIA"
    if any(k in up for k in ["INTERACCION", "INTERACCIÓN"]):
        return "INTERACCION"
    if any(k in up for k in ["INDICACION", "INDICACIÓN", "TERAPEUTICA"]):
        return "INDICACION"
    if any(k in up for k in ["FICHA TECNICA", "FICHA TÉCNICA", "RCP", "RESUMEN DE CARACTER"]):
        return "FICHA TECNICA"
    return "GENERAL"


def _analizar_heuristico(titulo: str, producto: str, texto: str) -> dict:
    tipo = _tipo_desde_titulo(titulo, producto)
    if tipo == "GENERAL" and texto:
        tipo = _tipo_desde_titulo(texto[:500], "")

    up = texto.upper() if texto else ""

    # Principio activo — buscar después de "principio activo:" o usar producto
    ifa = ""
    m = re.search(r"[Pp]rincipio\s+[Aa]ctivo[:\s]+([^\n]{3,60})", texto or "")
    if m:
        ifa = m.group(1).strip()

    # Titular
    titular = ""
    m2 = re.search(r"[Tt]itular[:\s]+([^\n]{3,80})", texto or "")
    if m2:
        titular = m2.group(1).strip()

    accion = {
        "SUSPENSION":           "Retirar producto del mercado y comunicar a DIGEMID en 5 días hábiles",
        "CANCELACION":          "Retirar definitivamente del mercado; informar en 5 días hábiles",
        "REACCION ADVERSA":     "Actualizar FT/rotulado y reportar en 15 días hábiles",
        "CONTRAINDICACION":     "Actualizar ficha técnica y material de prescripción en 30 días hábiles",
        "ADVERTENCIA":          "Actualizar ficha técnica y notificar a prescriptores en 30 días hábiles",
        "ACTUALIZACION SEGURIDAD": "Evaluar impacto y comunicar posición a DIGEMID en 15 días hábiles",
        "ROTULADO":             "Implementar cambio de rotulado en 90 días hábiles",
        "FICHA TECNICA":        "Actualizar ficha técnica en 60 días hábiles",
        "POSOLOGIA":            "Actualizar FT y material de prescripción en 60 días hábiles",
        "INTERACCION":          "Actualizar ficha técnica en 30 días hábiles",
        "INDICACION":           "Tramitar modificación de indicaciones ante DIGEMID",
        "INSCRIPCION":          "Verificar vigencia y condiciones del registro sanitario",
        "GENERAL":              "Revisar comunicado oficial y actuar según instrucciones DIGEMID",
    }.get(tipo, "Revisar comunicado oficial")

    resumen = ""
    for pat in [r"[Ss]e comunica[^\n.]{10,200}[.]", r"[Ss]e (informa|notifica)[^\n.]{10,200}[.]"]:
        m3 = re.search(pat, texto or "")
        if m3:
            resumen = m3.group(0)[:220]
            break

    return {"tipo_modificacion": tipo, "principio_activo": ifa,
            "titular_rs": titular, "accion_requerida": accion, "resumen": resumen}


# ══════════════════════════════════════════════════════════════════════════════
# PDF
# ══════════════════════════════════════════════════════════════════════════════
def _obtener_pdf_url(url_post: str) -> tuple[str | None, str]:
    """Devuelve (pdf_url, subtitulo). El subtítulo (h3 dentro de entry-content)
    suele contener el texto real del informe, ej. 'Informe sobre seguridad de
    los productos farmacéuticos que contienen Caspofungina' — mucho más
    informativo que el título genérico 'MODIFICACIONES N° 08 - 2026'."""
    r = _get_con_reintentos(url_post, timeout=20)
    if not r or r.status_code != 200:
        return None, ""
    try:
        soup = BeautifulSoup(r.text, "html.parser")
        entry = soup.find("div", class_="entry-content") or soup.find("div", class_=re.compile(r"content"))
        if not entry:
            return None, ""
        h3 = entry.find("h3")
        subtitulo = h3.get_text(strip=True) if h3 else ""
        link = entry.find("a", href=re.compile(r"\.pdf$", re.I))
        if link:
            href = link["href"]
            return (href if href.startswith("http") else f"{BASE_URL}{href}"), subtitulo
        embed = entry.find("embed", src=re.compile(r"\.pdf$", re.I))
        if embed:
            src = embed.get("src", "")
            return (src if src.startswith("http") else f"{BASE_URL}{src}"), subtitulo
        return None, subtitulo
    except Exception:
        return None, ""


def _descargar_pdf(url: str, reintentos: int = 3) -> bytes | None:
    hdrs = {**_HEADERS, "Accept": "application/pdf,*/*"}
    for i in range(1, reintentos + 1):
        try:
            r = _session.get(url, headers=hdrs, timeout=30)
            if r.status_code == 200:
                return r.content
            elif r.status_code == 429:
                time.sleep(PDF_DELAY * i)
            else:
                return None
        except Exception:
            time.sleep(PDF_DELAY)
    return None


def _extraer_texto(pdf_bytes: bytes) -> str:
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        return "\n".join(p.get_text() for p in doc).strip()
    except Exception:
        return ""


# ══════════════════════════════════════════════════════════════════════════════
# ENRIQUECIMIENTO
# ══════════════════════════════════════════════════════════════════════════════
def _texto_clasificacion(item: dict) -> str:
    """Título oficial + subtítulo real del post (ej. 'Informe sobre seguridad...')
    concatenados — la señal más fuerte disponible antes de tocar el PDF."""
    return f"{item.get('n_modificacion','')} {item.get('subtitulo','')}".strip()


def _enriquecer(item: dict, analizar_pdfs: bool = True) -> dict:
    url = item.get("url")
    if not url or not analizar_pdfs:
        tipo = _tipo_desde_titulo(_texto_clasificacion(item), item.get("producto", "") or "")
        item.update({
            "tipo_modificacion": tipo,
            "urgencia":          URGENCIA_MAP.get(tipo, "INFORMATIVA"),
            "indicador_tiempos": PLAZO_MAP.get(tipo, "—"),
            "base_legal":        BASE_MAP.get(tipo, "—"),
            "principio_activo":  "",
            "titular_rs":        "",
            "accion_requerida":  "Ver comunicado oficial",
            "resumen":           "",
            "pdf_url":           None,
            "motor_analisis":    "Título (sin PDF)",
        })
        return _forzar_textos_seguridad(item)

    pdf_url, subtitulo = _obtener_pdf_url(url)
    item["pdf_url"] = pdf_url
    item["subtitulo"] = subtitulo

    if not pdf_url:
        tipo = _tipo_desde_titulo(_texto_clasificacion(item), item.get("producto", "") or "")
        item.update({
            "tipo_modificacion": tipo, "urgencia": URGENCIA_MAP.get(tipo, "INFORMATIVA"),
            "indicador_tiempos": PLAZO_MAP.get(tipo, "—"), "base_legal": BASE_MAP.get(tipo, "—"),
            "principio_activo": "", "titular_rs": "", "accion_requerida": "Sin PDF disponible",
            "resumen": "", "motor_analisis": "Sin PDF",
        })
        return _forzar_textos_seguridad(item)

    pdf_bytes = _descargar_pdf(pdf_url)
    if not pdf_bytes:
        tipo = _tipo_desde_titulo(_texto_clasificacion(item), item.get("producto", "") or "")
        item.update({
            "tipo_modificacion": tipo, "urgencia": URGENCIA_MAP.get(tipo, "INFORMATIVA"),
            "indicador_tiempos": PLAZO_MAP.get(tipo, "—"), "base_legal": BASE_MAP.get(tipo, "—"),
            "principio_activo": "", "titular_rs": "", "accion_requerida": "Error descarga PDF",
            "resumen": "", "motor_analisis": "Error PDF",
        })
        return _forzar_textos_seguridad(item)

    texto = _extraer_texto(pdf_bytes)

    if len(texto) < 50:
        tipo = _tipo_desde_titulo(_texto_clasificacion(item), item.get("producto", "") or "")
        resultado = {"tipo_modificacion": tipo, "principio_activo": "",
                     "titular_rs": "", "accion_requerida": "PDF escaneado (imagen)", "resumen": ""}
        motor = "PDF escaneado"
    else:
        resultado = _analizar_claude(texto)
        if resultado:
            motor = "Claude API"
        else:
            resultado = _analizar_heuristico(_texto_clasificacion(item), item.get("producto","") or "", texto)
            motor = "Heurístico"

    tipo = resultado.get("tipo_modificacion", "GENERAL")
    item["tipo_modificacion"] = tipo
    item["urgencia"]          = URGENCIA_MAP.get(tipo, "INFORMATIVA")
    item["indicador_tiempos"] = PLAZO_MAP.get(tipo, "—")
    item["base_legal"]        = BASE_MAP.get(tipo, "—")
    item["principio_activo"]  = resultado.get("principio_activo", "")
    item["titular_rs"]        = resultado.get("titular_rs", "")
    item["accion_requerida"]  = resultado.get("accion_requerida", "")
    item["resumen"]           = resultado.get("resumen", "")
    item["motor_analisis"]    = motor
    return _forzar_textos_seguridad(item)


def _forzar_textos_seguridad(item: dict) -> dict:
    """Sobrescribe Acción Requerida / Indicador de Tiempos / Base Legal con los
    textos estandarizados cuando la modificatoria es de tipo seguridad.
    Sin esto, Claude API devuelve una acción redactada libremente y cada fila
    del Excel diría algo distinto."""
    if item.get("tipo_modificacion") in TIPOS_SEGURIDAD:
        item["accion_requerida"]  = ACCION_SEGURIDAD
        item["indicador_tiempos"] = TIEMPOS_SEGURIDAD
        item["base_legal"]        = BASE_LEGAL_SEGURIDAD
    return item


# ══════════════════════════════════════════════════════════════════════════════
# PARSEO DE PÁGINAS
# ══════════════════════════════════════════════════════════════════════════════
def _parsear_posts(soup: BeautifulSoup) -> list[dict]:
    items = []
    for article in soup.find_all("article", class_=re.compile(r"\bpost\b")):
        titulo_tag = article.find("h2", class_="entry-title") or article.find("h3", class_="entry-title")
        if not titulo_tag:
            continue
        titulo = titulo_tag.get_text(strip=True)
        if "MODIFICACION" not in titulo.upper() and "MODIFICACIÓN" not in titulo.upper():
            continue
        link_tag = titulo_tag.find("a")
        link = link_tag["href"] if link_tag and link_tag.get("href") else None
        time_tag = article.find("time")
        fecha_pub = None
        if time_tag and time_tag.get("datetime"):
            try:
                fecha_pub = datetime.strptime(time_tag["datetime"][:10], "%Y-%m-%d").date()
            except ValueError:
                pass
        excerpt = article.find("p", class_=re.compile(r"excerpt|summary", re.I)) or article.find("p")
        producto = excerpt.get_text(strip=True) if excerpt else None
        items.append({
            "n_modificacion": titulo,
            "producto":       producto,
            "fecha_publicacion": fecha_pub,
            "url":            link,
            "fecha_captura":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
    return items


def _total_paginas(soup: BeautifulSoup) -> int:
    pag = soup.find("div", class_="pagination") or soup.find("nav", class_=re.compile(r"pag"))
    if not pag:
        return 1
    nums = [int(a.get_text(strip=True))
            for a in pag.find_all("a", class_="page-numbers")
            if a.get_text(strip=True).isdigit()]
    return max(nums) if nums else 1


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
def scrapear_modificaciones(max_paginas: int = 2,
                             analizar_pdfs: bool = True,
                             solo_hoy: bool = False,
                             delay_paginas: float = 1.5) -> pd.DataFrame:
    resp = _get_con_reintentos(LIST_URL, timeout=30)
    if not resp:
        raise RuntimeError(f"No se pudo obtener el listado de modificatorias tras varios intentos: {LIST_URL}")
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    total = _total_paginas(soup)
    if max_paginas is not None:
        total = min(total, max_paginas)

    print(f"📋 Páginas a procesar: {total}")
    items = _parsear_posts(soup)
    print(f"  Página 1: {len(items)} modificatorias")

    for pag in range(2, total + 1):
        r = _get_con_reintentos(PAGE_URL.format(page=pag), timeout=30)
        if not r:
            print(f"  ERROR página {pag}: sin respuesta tras varios reintentos")
            time.sleep(delay_paginas)
            continue
        try:
            r.raise_for_status()
            nuevos = _parsear_posts(BeautifulSoup(r.text, "html.parser"))
            items.extend(nuevos)
            print(f"  Página {pag}/{total}: {len(nuevos)} modificatorias")
        except Exception as e:
            print(f"  ERROR página {pag}: {e}")
        time.sleep(delay_paginas)

    if solo_hoy:
        hoy = datetime.now().date()
        items = [i for i in items if i.get("fecha_publicacion") == hoy]
        print(f"  Filtro 'solo_hoy': {len(items)} publicaciones de hoy")

    motor_txt = "Claude API" if os.environ.get("ANTHROPIC_API_KEY") else "Heurístico"
    print(f"\n🔍 Analizando {len(items)} modificatorias [{motor_txt}]...")
    for idx, item in enumerate(items, 1):
        print(f"  [{idx:>3}/{len(items)}] {item['n_modificacion'][:70]}...", end=" ", flush=True)
        _enriquecer(item, analizar_pdfs)
        urgencia = item.get("urgencia", "—")
        emoji = {"INMEDIATA": "🔴", "PREVENTIVA": "🟡", "INFORMATIVA": "🔵"}.get(urgencia, "⚪")
        print(f"{emoji} [{item.get('tipo_modificacion','?'):22}] {item.get('motor_analisis','?')}")
        if analizar_pdfs:
            time.sleep(PDF_DELAY)

    df = pd.DataFrame(items)
    print(f"\n✅ Total modificatorias: {len(df)}")
    if "urgencia" in df.columns:
        print(df["urgencia"].value_counts().to_string())

    FALLOS_PDF = {"Sin PDF", "Error descarga PDF", "PDF escaneado"}
    if "motor_analisis" in df.columns:
        n_sin_pdf = int(df["motor_analisis"].isin(FALLOS_PDF).sum())
        if n_sin_pdf > 0:
            print(f"⚠️  {n_sin_pdf}/{len(df)} modificatoria(s) sin texto de PDF disponible "
                  f"(clasificadas solo por título/subtítulo — posible bloqueo o PDF caído)")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTAR EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def exportar_excel(df: pd.DataFrame, ruta: str, titulo: str = TITULO_REPORTE):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLS = [
        ("N° Modificación",      "n_modificacion",    45),
        ("Producto / IFA",        "producto",          30),
        ("Principio Activo",      "principio_activo",  25),
        ("Titular RS",            "titular_rs",        30),
        ("Fecha Publicación",     "fecha_publicacion", 14),
        ("Tipo de Modificación",  "tipo_modificacion", 22),
        ("Urgencia",              "urgencia",          12),
        ("Acción Requerida",      "accion_requerida",  45),
        ("⏱ Indicador Tiempos",  "indicador_tiempos", 50),
        ("Base Legal",            "base_legal",        28),
        ("Resumen IA",            "resumen",           55),
        ("Motor Análisis",        "motor_analisis",    14),
        ("URL Publicación",       "url",               50),
        ("URL PDF",               "pdf_url",           50),
        ("Fecha Captura",         "fecha_captura",     16),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modificatorias DIGEMID"

    thin   = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    d_font = Font(name="Arial", size=9)

    # ── Cabecera de reporte ConkoSafe (filas 1-2) ─────────────────────────────
    last_col = get_column_letter(len(COLS))
    if "fecha_publicacion" in df.columns and len(df):
        _f = pd.to_datetime(df["fecha_publicacion"], errors="coerce").dropna()
        desde = _f.min().strftime("%d/%m/%Y") if len(_f) else "—"
        hasta = _f.max().strftime("%d/%m/%Y") if len(_f) else "—"
    else:
        desde = hasta = "—"

    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = f"{titulo}  |  Desde {desde} hasta {hasta}"
    t.font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    for ci in range(1, len(COLS) + 1):
        ws.cell(row=1, column=ci).fill = h_fill
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last_col}2")
    s = ws["A2"]
    s.value = f"Generado: {datetime.now():%d/%m/%Y %H:%M}  |  Registros: {len(df)}"
    s.font = Font(italic=True, size=10, color="595959", name="Arial")
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Encabezados de columna (fila 3) ───────────────────────────────────────
    for ci, (label, _, width) in enumerate(COLS, 1):
        c = ws.cell(row=3, column=ci, value=label)
        c.font = h_font; c.fill = h_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 28

    URGENCIA_BG    = {"INMEDIATA": "FCE4D6", "PREVENTIVA": "FFF2CC", "INFORMATIVA": "EBF3FB"}
    URGENCIA_COLOR = {"INMEDIATA": "C00000", "PREVENTIVA": "ED7D31", "INFORMATIVA": "2E75B6"}
    TIEMPOS_FILL   = PatternFill("solid", start_color="E2EFDA")  # verde claro

    for ri, row in enumerate(df.itertuples(index=False), 4):
        urgencia = getattr(row, "urgencia", None) or "INFORMATIVA"
        row_fill = PatternFill("solid", start_color=URGENCIA_BG.get(urgencia, "F2F2F2"))
        for ci, (_, field, _) in enumerate(COLS, 1):
            val = getattr(row, field, None)
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = d_font; c.border = border
            c.alignment = Alignment(wrap_text=True, vertical="center")
            # Columna indicador tiempos → verde
            if field == "indicador_tiempos":
                c.fill = TIEMPOS_FILL
                c.font = Font(name="Arial", size=9, color="375623")
            else:
                c.fill = row_fill
            # Hipervínculos
            if field in ("url", "pdf_url") and val:
                c.hyperlink = str(val)
                c.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            # Urgencia en negrita con color
            if field == "urgencia" and val:
                c.font = Font(name="Arial", size=9, bold=True,
                              color=URGENCIA_COLOR.get(urgencia, "000000"))
                c.alignment = Alignment(horizontal="center", vertical="center")
            if field == "accion_requerida":
                c.font = Font(name="Arial", size=9, bold=True)
        ws.row_dimensions[ri].height = 55

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col}{len(df)+3}"

    # ── Hoja 2: Resumen Diario ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen Diario")
    ws2["A1"] = titulo
    ws2["A1"].font = Font(bold=True, size=13, name="Arial", color="1F4E79")
    ws2["A3"] = "Fecha captura:"; ws2["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws2["A4"] = "Total modificatorias:"; ws2["B4"] = len(df)

    ws2["A6"] = "Por Urgencia"; ws2["A6"].font = Font(bold=True, name="Arial", color="1F4E79")
    if "urgencia" in df.columns:
        for i, (k, v) in enumerate(df["urgencia"].value_counts().items(), 7):
            ws2[f"A{i}"] = k; ws2[f"B{i}"] = v

    ws2["D6"] = "Por Tipo de Modificación"; ws2["D6"].font = Font(bold=True, name="Arial", color="1F4E79")
    if "tipo_modificacion" in df.columns:
        for i, (k, v) in enumerate(df["tipo_modificacion"].value_counts().items(), 7):
            ws2[f"D{i}"] = k; ws2[f"E{i}"] = v

    ws2["G6"] = "Productos más frecuentes"; ws2["G6"].font = Font(bold=True, name="Arial", color="1F4E79")
    if "producto" in df.columns:
        for i, (k, v) in enumerate(df["producto"].dropna().value_counts().head(8).items(), 7):
            ws2[f"G{i}"] = str(k)[:40]; ws2[f"H{i}"] = v

    for col in ["A","B","D","E","G","H"]:
        ws2.column_dimensions[col].width = 42

    wb.save(ruta)
    print(f"💾 Excel guardado: {ruta}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    motor = "Claude API" if os.environ.get("ANTHROPIC_API_KEY") else "Heurístico"
    print(f"Motor: {motor}")
    print(f"URL:   {LIST_URL}\n")

    # Scrapear suficientes páginas para encontrar 10 de ACTUALIZACION SEGURIDAD
    df = scrapear_modificaciones(max_paginas=2, analizar_pdfs=True, solo_hoy=False)

    # Filtrar solo ACTUALIZACION SEGURIDAD y tomar las 10 más recientes
    if "tipo_modificacion" in df.columns:
        df = df[df["tipo_modificacion"].isin(TIPOS_SEGURIDAD)]
        if "fecha_publicacion" in df.columns:
            df = df.sort_values("fecha_publicacion", ascending=False)
        df = df.head(10).reset_index(drop=True)
    print(f"📌 Filtradas: {len(df)} modificatorias de seguridad")

    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    out_dir = os.path.join(os.getcwd(), "output")
    os.makedirs(out_dir, exist_ok=True)
    ruta = os.path.join(out_dir, f"modificatorias_digemid_{fecha}.xlsx")
    exportar_excel(df, ruta)
    print(f"\nArchivo listo: {ruta}")
